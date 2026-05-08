import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
import os
import numpy as np
from .logger import get_logger
from openpyxl.drawing.image import Image as OpenpyxlImage
from datetime import date

logger = get_logger(__name__)

def generate_excel_report(
    filename: str,
    params: dict,
    metrics: dict,
    benchmark: list
):
    """
    Genera un reporte en Excel de las simulaciones Monte Carlo (BSM vs Heston).
    
    Args:
        filename (str): Ruta de salida del archivo Excel (.xlsx).
        params (dict): Diccionario con parametros de la simulacion y modelos.
        metrics_bsm (dict): Metricas calculadas para BSM.
        metrics_heston (dict): Metricas calculadas para Heston.
        sample_trajectories (np.ndarray): Array 2D (dias x num_trayectorias) con trayectorias de muestra.
    """
    wb = openpyxl.Workbook()
    
    # Estilos basicos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    border_thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def format_header(sheet, row, cols):
        for col in cols:
            cell = sheet.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = border_thin


    def write_metrics_sheet(ws, title_sheet, metrics_dict):


        metrics_keys = [
        ("VaR 95%", "VaR_95"),
        ("CVaR 95%", "CVaR_95"),
        ("Probabilidad de Perdida", "prob_loss"),
        ("Probabilidad de Ganancia > 20%", "prob_gain_20"),
        ("MSE Precio", "mse_price"),
        ("MAE Precio", "mae_price"),
        ("MSE Crecimiento", "mse_growth"),
        ("MAE Crecimiento", "mae_growth"),
        ("KS Test (p-value)", "ks_test_p"),
        ("KS test (distribucion)", "ks_test_d"),
        ("Coverage IC 90%", "coverage")
        ]
        ws.title = title_sheet
        ws.cell(row=1, column=1, value="Metrica")
        ws.cell(row=1, column=2, value="Valor")
        format_header(ws, 1, [1, 2])
        
        for row_idx, (label, key) in enumerate(metrics_keys, start=2):
            ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
            val_cell = ws.cell(row=row_idx, column=2, value=metrics_dict.get(key, "N/A"))
            ws.cell(row=row_idx, column=1).border = border_thin
            val_cell.border = border_thin
            # Number formatting
            if "Probabilidad" in label or "Coverage" in label:
                val_cell.number_format = '0.00%'
            elif type(val_cell.value) in [float, int, np.float64]:
                val_cell.number_format = '0.0000'

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15


    params_data = [
        ("Parametros Generales", ""),
        ("Ticker", params.get("ticker", "N/A")),
        ("Fecha de Inicio", params.get("start_date", "N/A")),
        ("Fecha de Final", params.get("end_date", "N/A")),
        ("Fecha de Simulacion", params.get("simulation_date", "N/A")),
        ("N Simulaciones", params.get("n_simulaciones", 0)),
        ("", ""),
        ("Parametros del Activo", ""),
        ("Precio Inicial (S0)", params.get("precio_inicial", 0.0)),
        ("Rendimiento Esperado (mu)", params.get("mu", 0.0)),
        ("Volatilidad (sigma)", params.get("sigma", 0.0)),
        ("", ""),
        ("Parametros del modelo", ""),
        ("Id", params.get("Id", "")),
        ("Modelo",params.get("modelo", "")),
        ("Correlacion (rho)", params.get("rho", 0.0)),
        ("Tasa de Reversion (kappa)", params.get("kappa", 0.0)),
        ("Varianza a Largo Plazo (theta)", params.get("theta", 0.0)),
        ("Varianza Inicial (v0)", params.get("v0", 0.0)),
    ]




    # ==========================
    # HOJA 1 - Insights y Graficos
    # ==========================
    ws1 = wb.create_sheet(title="Insights_MC")
    
    # Adding the insight message
    var_95 = metrics.get('VaR_95', 0.0)
    cvar_95 = metrics.get('CVaR_95', 0.0)
    
    insigth_MC = f"Análisis Provisional: El modelo {params.get('modelo', 'N/A')} muestra un VaR(95%) de {var_95:.2%} y un CVaR(95%) de {cvar_95:.2%}. Esto indica el riesgo de caída esperado bajo las simulaciones de Monte Carlo."
    
    #Bajada
    ws1.cell(row=1,column=1,value = "Dashboard de Simulaciones Monte Carlo")
    ws1.cell(row=1,column=1).font = Font(bold=True, size=14)

    #columnas
    ws1.cell(row=3,column=2, value="Modelo")        
    ws1.cell(row=3,column=3,value ="Fecha")
    ws1.cell(row=3,column=4,value = "Ticker")
    ws1.cell(row=3,column=5,value = "N Simulaciones")
    ws1.cell(row=3,column=6,value = "MAE")
    ws1.cell(row=3,column=7,value = "Coverage")
    for idx_colum in range(2,8):
        ws1.cell(row=3,column=idx_colum).border = border_thin
        ws1.cell(row=3,column=idx_colum).font = Font(bold=True)
        ws1.cell(row=3,column=idx_colum).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    #Valores
    ws1.cell(row=4, column=2, value = f"{params.get('modelo', 'N/A')}V1" )
    ws1.cell(row=4,column=3,value = f"{date.today().strftime('%Y-%m-%d')}" )
    ws1.cell(row=4,column=4, value = f"{params.get('ticker', 'N/A')}" )
    ws1.cell(row=4,column=5,value = f"{params.get('n_simulaciones', 'N/A')}" )
    ws1.cell(row=4,column=6,value = f"{metrics.get('mae_price', 'N/A')}" )
    ws1.cell(row=4,column=7, value = f"{metrics.get('coverage', 'N/A')}" )

    #Bordes
    for col_idx in range(2,8):
        ws1.cell(row=3, column=col_idx).border = border_thin
        ws1.cell(row=4, column=col_idx).border = border_thin

    #fondo de las columnas    

    
    # Adding the image
    img_path = f"Reporte Monte Carlo {params.get('modelo')}.png"
    if os.path.exists(img_path):
        try:
            img = OpenpyxlImage(img_path)
 
            img.width = int(img.width * 0.7)
            img.height = int(img.height * 0.7)
            
            # Insertar la primera imagen en B5
            ws1.add_image(img, 'A9')
            
            # --- AGREGAR OTRA IMAGEN AL LADO ---
            img2_path = "Reporte Visual del Activo.png" 
            if os.path.exists(img2_path):
                img2 = OpenpyxlImage(img2_path)
                img2.width = int(img2.width * 0.5)
                img2.height = int(img2.height * 0.5)
                ws1.add_image(img2, 'M9')
                
        except ImportError:
            logger.warning("Pillow no esta instalado. No se pudo insertar la imagen.")
            ws1.cell(row=6, column=2, value="Error: Pillow no esta instalado. Instale Pillow para ver la imagen.")
    else:
        logger.warning(f"Imagen no encontrada: {img_path}")
        ws1.cell(row=6, column=2, value=f"Imagen no encontrada: {img_path}")

    ws1.cell(row=7, column=2, value="Comentario / Insight:")
    ws1.cell(row=7, column=2).font = Font(bold=True)
    ws1.cell(row=7, column=2, value=insigth_MC)



    # ==========================
    # HOJA 2 - Metricas
    # ==========================
    ws2 = wb.create_sheet()
    write_metrics_sheet(ws2, "Metricas", metrics)

    # ==========================
    # HOJA 3 - Modelos Historicos
    # ==========================
    ws3 = wb.create_sheet()
    ws3.title = "Modelos Historicos"

    # Encabezados de la tabla
    headers = ["Id", "Modelo", "N_simulaciones", "Tiempo Ejecución", "MAE", "MSE", "Coverage", "Parámetros"]
    for col_num, header in enumerate(headers, 1):
        cell = ws3.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = border_thin

    current_id = params.get("Id", "")
    
    # Manejo de benchmark en caso de ser lista o diccionario
    if isinstance(benchmark, list):
        benchmark_list = benchmark
    elif isinstance(benchmark, dict) and "Id" in benchmark:
        benchmark_list = [benchmark]
    else:
        benchmark_list = []

    for row_idx, result in enumerate(benchmark_list, start=2):
        row_id = result.get("Id", "")
        modelo_nombre = result.get("modelo", "")
        n_sim = result.get("N_simulaciones", "")
        t_ejec = result.get("tiempo_ejecucion", "")
        mae = result.get("MAE", "")
        mse = result.get("MSE", "")
        coverage = result.get("Coverage", "")
        parametros = str(result.get("parametros", ""))

        is_current = (str(row_id) == str(current_id))
        font_style = Font(bold=True) if is_current else Font(bold=False)

        row_data = [row_id, modelo_nombre, n_sim, t_ejec, mae, mse, coverage, parametros]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_num, value=value)
            cell.font = font_style
            cell.border = border_thin
            if isinstance(value, (int, float, np.float64)) and not isinstance(value, bool):
                if headers[col_num-1] == "Coverage":
                    cell.number_format = '0.00%'
                else:
                    cell.number_format = '0.0000'

    # Ajustar anchos de columna
    for col_letter, width in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H'], [10, 15, 15, 18, 12, 12, 12, 50]):
        ws3.column_dimensions[col_letter].width = width


    # ==========================
    # HOJA 4 - Parametros
    # ==========================
    ws4 = wb.create_sheet(title = "Parametros")

    for row_idx, (key, value) in enumerate(params_data, start=1):
        ws4.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
        if value != "":
            ws4.cell(row=row_idx, column=2, value=value)
            ws4.cell(row=row_idx, column=1).border = border_thin
            ws4.cell(row=row_idx, column=2).border = border_thin
        else:
            ws4.cell(row=row_idx, column=1).fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    ws4.column_dimensions['A'].width = 30
    ws4.column_dimensions['B'].width = 20





    # ==========================
    # Guardar
    # ==========================
    wb.save(filename)
    logger.info(f"Reporte de Excel guardado exitosamente en: {filename}")


